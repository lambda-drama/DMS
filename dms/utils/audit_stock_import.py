"""Create a draft Stock Reconciliation from an Audit Report Stock workbook."""

from __future__ import annotations

import os
from dataclasses import dataclass

import frappe
from frappe import _
from frappe.utils import cint, flt

from dms.dealer_management_system.utils.stock_operations import create_dms_stock_reconciliation
from dms.utils.inventory_stock_import import (
	DEFAULT_STOCK_UOM,
	INVENTORY_WAREHOUSE,
	_default_item_company,
	_norm_header,
	_number,
	_resolve_file_path,
	_text,
	_warehouse_company,
)

_HEADER_ALIASES = {
	"part_no": {"partno", "partnumber", "partno."},
	"part_name": {"partname"},
	"qty": {"qty", "qtypcs", "qty(pcs)", "qtypcs)"},
	"physical_stock": {"physicalstock", "physicalqty", "physical"},
	"location": {"location"},
}

_VALUATION_HEADER_ALIASES = {
	"part_no": {"partno", "partnumber", "partno."},
	"total_unit_price_with_tt": {"totalunitpricewithtt"},
}

AUDIT_PURPOSES = ("Opening Stock", "Stock Reconciliation")


@dataclass(slots=True)
class AuditStockRow:
	part_no: str
	part_name: str
	qty: float
	location: str


def create_audit_stock_reconciliation_file_url(
	file_url: str,
	valuation_file_url: str | None = None,
	posting_date: str | None = None,
	submit: int | bool = 0,
	purpose: str | None = None,
	expense_account: str | None = None,
	warehouse: str = INVENTORY_WAREHOUSE,
) -> dict:
	return create_audit_stock_reconciliation_from_workbook(
		_resolve_file_path(file_url),
		valuation_file_path=_resolve_file_path(valuation_file_url) if valuation_file_url else None,
		posting_date=posting_date,
		submit=submit,
		purpose=purpose,
		expense_account=expense_account,
		warehouse=warehouse,
	)


def create_audit_stock_reconciliation_from_workbook(
	file_path: str,
	valuation_file_path: str | None = None,
	posting_date: str | None = None,
	submit: int | bool = 0,
	purpose: str | None = None,
	expense_account: str | None = None,
	warehouse: str = INVENTORY_WAREHOUSE,
) -> dict:
	purpose = (purpose or "Stock Reconciliation").strip()
	if purpose not in AUDIT_PURPOSES:
		frappe.throw(_("Purpose must be Opening Stock or Stock Reconciliation."))

	warehouse = (warehouse or INVENTORY_WAREHOUSE).strip()
	company = _warehouse_company(warehouse)
	expense_account = _validate_difference_account(expense_account, purpose, company)

	rows = _read_audit_stock_rows(file_path)
	valuation_rates = _read_valuation_rates(valuation_file_path) if valuation_file_path else {}

	items, duplicate_rows_merged = _aggregate_audit_rows(rows)
	summary = {
		"items_created": 0,
		"items_reused": 0,
		"spare_parts_created": 0,
		"spare_parts_updated": 0,
		"locations_updated": 0,
		"rates_applied": 0,
		"rates_missing": 0,
		"skipped": [],
	}

	recon_items: list[dict] = []
	for row in items:
		item_code, skip_reason = _ensure_item_and_spare_part(row, summary)
		if skip_reason:
			summary["skipped"].append({"part_no": row.part_no, "reason": skip_reason})
			continue
		if row.location:
			if _update_spare_part_bin_location(item_code, row.location):
				summary["locations_updated"] += 1
		line = {"item_code": item_code, "qty": max(flt(row.qty), 0)}
		rate = flt(valuation_rates.get(row.part_no) or valuation_rates.get(item_code) or 0)
		if rate > 0:
			line["valuation_rate"] = rate
			summary["rates_applied"] += 1
		else:
			summary["rates_missing"] += 1
		recon_items.append(line)

	if not recon_items:
		frappe.throw(_("No matching spare part items were found in the audit workbook."))

	result = create_dms_stock_reconciliation(
		{
			"company": company,
			"warehouse": warehouse,
			"posting_date": posting_date or frappe.utils.today(),
			"submit": cint(submit),
			"purpose": purpose,
			"expense_account": expense_account,
			"remarks": _("Audit report {0} from workbook").format(purpose.lower()),
			"items": recon_items,
		}
	)
	frappe.db.commit()
	return {
		**result,
		"warehouse": warehouse,
		"company": company,
		"purpose": purpose,
		"expense_account": expense_account,
		"rows_processed": len(rows),
		"unique_items": len(recon_items),
		"duplicate_rows_merged": duplicate_rows_merged,
		"skipped_count": len(summary["skipped"]),
		"skipped": summary["skipped"][:50],
		"items_created": summary["items_created"],
		"items_reused": summary["items_reused"],
		"spare_parts_created": summary["spare_parts_created"],
		"spare_parts_updated": summary["spare_parts_updated"],
		"locations_updated": summary["locations_updated"],
		"rates_applied": summary["rates_applied"],
		"rates_missing": summary["rates_missing"],
	}


def _read_audit_stock_rows(file_path: str) -> list[AuditStockRow]:
	try:
		import openpyxl
	except ImportError as exc:
		raise ImportError(_("Excel import requires openpyxl in the bench environment.")) from exc

	if not os.path.isfile(file_path):
		frappe.throw(_("File not found: {0}").format(file_path))

	wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
	try:
		ws = wb[wb.sheetnames[0]]
		raw_rows = list(ws.iter_rows(values_only=True))
	finally:
		wb.close()

	if not raw_rows:
		return []

	header_map = _header_map_with_aliases(raw_rows[0], _HEADER_ALIASES)
	required = ("part_no", "part_name")
	missing = [field for field in required if field not in header_map]
	if missing:
		frappe.throw(_("Audit workbook is missing columns: {0}").format(", ".join(missing)))
	if "physical_stock" not in header_map and "qty" not in header_map:
		frappe.throw(_("Audit workbook is missing a Physical Stock or QTY column."))
	if "location" not in header_map:
		frappe.throw(_("Audit workbook is missing column: location"))

	rows: list[AuditStockRow] = []
	for raw in raw_rows[1:]:
		part_no = _normalize_part_no(raw[header_map["part_no"]] if len(raw) > header_map["part_no"] else None)
		part_name = _text(raw[header_map["part_name"]] if len(raw) > header_map["part_name"] else None)
		if not part_no or not part_name:
			continue

		physical = None
		if "physical_stock" in header_map and len(raw) > header_map["physical_stock"]:
			physical = raw[header_map["physical_stock"]]
		system_qty = None
		if "qty" in header_map and len(raw) > header_map["qty"]:
			system_qty = raw[header_map["qty"]]
		qty = _number(physical if physical is not None else system_qty)
		location = _text(raw[header_map["location"]] if len(raw) > header_map["location"] else None)
		rows.append(
			AuditStockRow(
				part_no=part_no,
				part_name=part_name[:140],
				qty=qty,
				location=location,
			)
		)

	return rows


def _read_valuation_rates(file_path: str) -> dict[str, float]:
	try:
		import openpyxl
	except ImportError as exc:
		raise ImportError(_("Excel import requires openpyxl in the bench environment.")) from exc

	if not os.path.isfile(file_path):
		frappe.throw(_("Valuation file not found: {0}").format(file_path))

	wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
	try:
		ws = wb[wb.sheetnames[0]]
		raw_rows = list(ws.iter_rows(values_only=True))
	finally:
		wb.close()

	if not raw_rows:
		return {}

	header_map = _header_map_with_aliases(raw_rows[0], _VALUATION_HEADER_ALIASES)
	if "part_no" not in header_map or "total_unit_price_with_tt" not in header_map:
		frappe.throw(
			_("Valuation workbook must include Part No and Total Unit Price with T/T columns.")
		)

	rates: dict[str, float] = {}
	for raw in raw_rows[1:]:
		part_no = _normalize_part_no(raw[header_map["part_no"]] if len(raw) > header_map["part_no"] else None)
		if not part_no:
			continue
		rate = _number(
			raw[header_map["total_unit_price_with_tt"]]
			if len(raw) > header_map["total_unit_price_with_tt"]
			else None
		)
		if rate > 0:
			rates[part_no] = rate
	return rates


def _header_map_with_aliases(header_row: tuple | list, aliases: dict[str, set[str]]) -> dict[str, int]:
	out: dict[str, int] = {}
	for idx, raw in enumerate(header_row or []):
		key = _norm_header(raw)
		if not key:
			continue
		for fieldname, field_aliases in aliases.items():
			if key in field_aliases:
				out.setdefault(fieldname, idx)
	return out


def _normalize_part_no(value) -> str:
	return _text(value).rstrip(".")


def _aggregate_audit_rows(rows: list[AuditStockRow]) -> tuple[list[AuditStockRow], int]:
	aggregated: dict[str, AuditStockRow] = {}
	duplicate_rows_merged = 0
	for row in rows:
		existing = aggregated.get(row.part_no)
		if not existing:
			aggregated[row.part_no] = AuditStockRow(
				part_no=row.part_no,
				part_name=row.part_name,
				qty=flt(row.qty),
				location=row.location,
			)
			continue
		duplicate_rows_merged += 1
		existing.qty += flt(row.qty)
		if row.location:
			existing.location = row.location
	return list(aggregated.values()), duplicate_rows_merged


def _ensure_item_and_spare_part(row: AuditStockRow, summary: dict) -> tuple[str | None, str | None]:
	item_code = _resolve_item_code(row.part_no)
	if item_code:
		if not cint(frappe.db.get_value("Item", item_code, "is_stock_item")):
			return None, _("Item {0} is not a stock item").format(item_code)
		summary["items_reused"] += 1
	else:
		item_code = _create_item(row)
		summary["items_created"] += 1

	spare_name = _find_spare_part(item_code, row.part_no)
	if spare_name:
		summary["spare_parts_updated"] += 1
	else:
		_create_spare_part(item_code, row)
		summary["spare_parts_created"] += 1
	return item_code, None


def _resolve_item_code(part_no: str) -> str | None:
	if frappe.db.exists("Item", part_no):
		return part_no

	spare_name = _find_spare_part(part_no, part_no)
	if spare_name:
		linked = (frappe.db.get_value("Spare Part", spare_name, "spare_part_item") or "").strip()
		if linked and frappe.db.exists("Item", linked):
			return linked
	return None


def _find_spare_part(item_code: str, part_no: str) -> str | None:
	return (
		frappe.db.get_value("Spare Part", {"spare_part_item": item_code}, "name")
		or frappe.db.get_value("Spare Part", {"oem_part_number": part_no}, "name")
		or frappe.db.get_value("Spare Part", {"oem_part_number": item_code}, "name")
		or (item_code if frappe.db.exists("Spare Part", item_code) else None)
		or (part_no if part_no != item_code and frappe.db.exists("Spare Part", part_no) else None)
	)


def _create_item(row: AuditStockRow) -> str:
	item_group = (frappe.db.get_single_value("DMS Settings", "default_item_group") or "").strip()
	if not item_group or not frappe.db.exists("Item Group", item_group):
		frappe.throw(_("Set Default Item Group on DMS Settings before creating missing audit items."))

	doc = frappe.get_doc(
		{
			"doctype": "Item",
			"item_code": row.part_no,
			"item_name": row.part_name,
			"description": row.part_name,
			"item_group": item_group,
			"stock_uom": DEFAULT_STOCK_UOM,
			"is_stock_item": 1,
			"is_sales_item": 1,
			"is_purchase_item": 1,
			"include_item_in_manufacturing": 0,
			"disabled": 0,
			"item_defaults": [{"company": _default_item_company()}],
		}
	)
	doc.insert(ignore_permissions=True)
	return doc.name


def _create_spare_part(item_code: str, row: AuditStockRow) -> str:
	doc = frappe.get_doc(
		{
			"doctype": "Spare Part",
			"spare_part_item": item_code,
			"oem_part_number": item_code,
			"manufacturer_part_number": item_code,
			"part_category": "Other",
			"part_type": "Genuine OEM",
			"bin_location": row.location or None,
		}
	)
	doc.insert(ignore_permissions=True)
	return doc.name


def _update_spare_part_bin_location(item_code: str, location: str) -> bool:
	location = (location or "").strip()
	if not location:
		return False
	spare_name = _find_spare_part(item_code, item_code)
	if not spare_name:
		return False
	current = (frappe.db.get_value("Spare Part", spare_name, "bin_location") or "").strip()
	if current == location:
		return False
	frappe.db.set_value("Spare Part", spare_name, "bin_location", location, update_modified=True)
	return True


def _validate_difference_account(account: str | None, purpose: str, company: str) -> str:
	account = (account or "").strip()
	if not account:
		frappe.throw(_("Difference Account is required."))
	if not frappe.db.exists("Account", account):
		frappe.throw(_("Account {0} was not found.").format(frappe.bold(account)))

	acc_company, is_group, report_type = frappe.db.get_value(
		"Account", account, ["company", "is_group", "report_type"]
	)
	if (acc_company or "").strip() != company:
		frappe.throw(_("Difference Account must belong to Company {0}.").format(frappe.bold(company)))
	if cint(is_group):
		frappe.throw(_("Difference Account cannot be a group account."))
	if purpose == "Opening Stock" and report_type == "Profit and Loss":
		frappe.throw(
			_(
				"Difference Account must be a Asset/Liability type account, since this Stock Reconciliation is an Opening Entry"
			)
		)
	return account
