# Copyright (c) 2026, Mania and contributors
"""Import Vehicle Service Packages from periodic maintenance price Excel workbooks."""

from __future__ import annotations

import os
import re

import frappe
from frappe import _
from frappe.utils import cint, flt, get_files_path

from dms.utils.frt_sheet_import import DEFAULT_BRAND, ensure_brand, ensure_vehicle_model


def _norm_key(value: str) -> str:
	return re.sub(r"[^A-Z0-9]", "", (value or "").upper())


SKIP_SHEETS = frozenset({"SUMMARY", "Summary"})
# Normalized price-workbook sheet name -> (model_name, model_code).
_KNOWN_SHEET_MODELS: dict[str, tuple[str, str]] = {
	_norm_key("X-50"): ("X50", "JX50"),
	_norm_key("X-70"): ("X70", "JX70"),
	_norm_key("X-70 PLUS"): ("X70 PLUS", "JX70P"),
	_norm_key("X-70 CDM"): ("X70 CDM", "JX70H"),
	_norm_key("X-90 PLUS"): ("X90 PLUS", "JX90P"),
	_norm_key("DASHING 1.5"): ("Dashing", "JD15"),
}
LABOUR_META_LABELS = frozenset(
	{
		"hours",
		"labour rate",
		"total labour",
		"less discount",
		"oil",
		"other consumables",
		"total oils & consumables",
		"parts",
		"total parts (retail)",
		"total parts (net)",
		"before disct",
		"after disct",
		"part no",
	}
)
PART_SECTION_STOP = frozenset(
	{
		"total parts (retail)",
		"less discount",
		"total parts (net)",
		"before disct",
		"after disct",
		"part no",
	}
)


def import_service_package_workbook(file_path: str) -> dict:
	try:
		import openpyxl
	except ImportError as exc:
		raise ImportError(_("Excel import requires openpyxl in the bench environment.")) from exc

	if not os.path.isfile(file_path):
		frappe.throw(_("File not found: {0}").format(file_path))

	wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
	interval_map = _parse_summary_intervals(wb)
	summary = {
		"packages_created": 0,
		"packages_updated": 0,
		"vehicle_models_created": 0,
		"sheets_processed": 0,
		"errors": [],
		"details": [],
	}

	for sheet_name in wb.sheetnames:
		if sheet_name in SKIP_SHEETS:
			continue
		try:
			result = _import_model_sheet(wb[sheet_name], sheet_name, interval_map)
			if not result:
				continue
			summary["sheets_processed"] += 1
			summary["packages_created"] += result["packages_created"]
			summary["packages_updated"] += result["packages_updated"]
			summary["vehicle_models_created"] += result.get("vehicle_model_created", 0)
			summary["details"].append(result)
		except Exception as exc:
			frappe.log_error(title=f"Service package import — {sheet_name}")
			summary["errors"].append({"sheet": sheet_name, "error": str(exc)})

	wb.close()
	frappe.db.commit()
	return summary


def import_service_package_file_url(file_url: str) -> dict:
	path = _resolve_file_path(file_url)
	return import_service_package_workbook(path)


def _resolve_file_path(file_url: str) -> str:
	file_url = (file_url or "").strip()
	if not file_url:
		frappe.throw(_("File URL is required"))

	if file_url.startswith("/private/files/"):
		return get_files_path(*file_url.replace("/private/files/", "").split("/"), is_private=1)

	if file_url.startswith("/files/"):
		return get_files_path(*file_url.replace("/files/", "").split("/"), is_private=0)

	if os.path.isabs(file_url) and os.path.isfile(file_url):
		return file_url

	frappe.throw(_("Could not resolve uploaded file path"))


def _import_model_sheet(ws, sheet_name: str, interval_map: dict) -> dict | None:
	rows = list(ws.iter_rows(values_only=True))
	if len(rows) < 10:
		return None

	header = rows[1] if len(rows) > 1 else ()
	sheet_label = _text(header[0] if header else sheet_name)
	vm = _resolve_vehicle_model(sheet_name, sheet_label)
	if not vm:
		return None

	model_name, model_code, vehicle_model, vehicle_model_created = vm
	packages = _parse_package_columns(header)
	if not packages:
		return None

	row_map = {_norm_label(rows[i][0]): i for i in range(len(rows)) if _text(rows[i][0])}
	hours_row = row_map.get("hours")
	labour_discount_row = _first_row_index(rows, "less discount", before_label="parts")
	before_discount_row = row_map.get("before disct")
	after_discount_row = row_map.get("after disct")
	parts_start = row_map.get("parts")
	part_catalog = _build_part_catalog(rows)

	stats = {
		"sheet": sheet_name,
		"vehicle_model": vehicle_model,
		"model_code": model_code,
		"model_name": model_name,
		"vehicle_model_created": vehicle_model_created,
		"packages_created": 0,
		"packages_updated": 0,
		"package_names": [],
	}

	for col_idx, package_id, interval_km in packages:
		hours = _cell_number(rows[hours_row][col_idx]) if hours_row is not None else 0
		labour_discount = 0.0
		if labour_discount_row is not None:
			labour_discount = _cell_number(rows[labour_discount_row][col_idx])

		interval_months = interval_map.get(interval_km, {}).get("months")
		description = interval_map.get(interval_km, {}).get("label") or _package_description(
			model_name, package_id, interval_km, interval_months
		)

		labor_operations = _build_labor_operations(vehicle_model, model_code, interval_km, hours)
		parts_included = _build_package_parts(rows, col_idx, parts_start, part_catalog)

		before_discount = (
			_cell_number(rows[before_discount_row][col_idx]) if before_discount_row is not None else 0
		)
		after_discount = (
			_cell_number(rows[after_discount_row][col_idx]) if after_discount_row is not None else 0
		)
		total_amount = after_discount or before_discount

		package_name = _package_name(model_code, package_id)
		created = upsert_service_package(
			package_name=package_name,
			package_id=package_id,
			vehicle_model=vehicle_model,
			model_name=model_name,
			description=description,
			interval_km=interval_km,
			interval_months=interval_months,
			total_labor_hours=hours,
			labour_discount_amount=labour_discount,
			before_discount=before_discount,
			after_discount=after_discount,
			total_amount=total_amount,
			labor_operations=labor_operations,
			parts_included=parts_included,
		)
		if created:
			stats["packages_created"] += 1
		else:
			stats["packages_updated"] += 1
		stats["package_names"].append(package_name)

	return stats


def upsert_service_package(
	package_name: str,
	package_id: str,
	vehicle_model: str,
	model_name: str,
	description: str,
	interval_km: int | None,
	interval_months: int | None,
	total_labor_hours: float,
	labour_discount_amount: float,
	before_discount: float,
	after_discount: float,
	total_amount: float,
	labor_operations: list[dict],
	parts_included: list[dict],
) -> bool:
	package_name = (package_name or "").strip()
	if not package_name:
		return False

	after_discount = flt(after_discount)
	before_discount = flt(before_discount)
	total_amount = flt(total_amount) or after_discount or before_discount

	values = {
		"package_name": package_name,
		"package_id": package_id,
		"vehicle_model": vehicle_model,
		"description": description or f"{model_name} · {package_id}",
		"interval_km": interval_km,
		"interval_months": interval_months,
		"total_labor_hours": total_labor_hours,
		"labour_discount_amount": labour_discount_amount,
		"before_discount": before_discount,
		"after_discount": after_discount,
		"total_amount": total_amount,
		"package_price": after_discount or total_amount,
		"is_active": 1,
		"applicable_vehicle_models": [{"vehicle_model": vehicle_model}],
		"labor_operations": labor_operations,
		"parts_included": parts_included,
	}

	existing = frappe.db.get_value("Vehicle Service Package", package_name, "name")
	if existing:
		doc = frappe.get_doc("Vehicle Service Package", existing)
		doc.set("labor_operations", [])
		doc.set("parts_included", [])
		doc.set("applicable_vehicle_models", [])
		doc.update(values)
		doc.save(ignore_permissions=True)
		return False

	doc = frappe.get_doc({"doctype": "Vehicle Service Package", **values})
	doc.insert(ignore_permissions=True)
	return True


def _build_labor_operations(
	vehicle_model: str, model_code: str, interval_km: int | None, hours: float
) -> list[dict]:
	vsi = _find_periodic_vsi(vehicle_model, model_code, interval_km, hours)
	if not vsi:
		return []

	return [
		{
			"labor_operation": vsi,
			"quantity": 1,
			"total_hours": hours or None,
		}
	]


def _find_periodic_vsi(
	vehicle_model: str, model_code: str, interval_km: int | None, hours: float
) -> str | None:
	pm_sub = _km_to_pm_sub_code(interval_km, hours)
	candidates = []
	if pm_sub is not None:
		candidates.append(f"{model_code}PM{pm_sub}")

	for code in candidates:
		name = frappe.db.get_value(
			"Vehicle Service Item",
			{"custom_service_code": code, "custom_vehicle_model": vehicle_model},
			"name",
		)
		if name:
			return name

	if pm_sub is not None:
		name = frappe.db.get_value(
			"Vehicle Service Item",
			{
				"custom_vehicle_model": vehicle_model,
				"custom_cat_code": "PM",
				"custom_sub_code": str(pm_sub),
			},
			"name",
		)
		if name:
			return name

	rows = frappe.get_all(
		"Vehicle Service Item",
		filters={"custom_vehicle_model": vehicle_model, "custom_cat_code": "PM"},
		fields=["name", "custom_sub_code", "custom_estimated_timehours", "service_item"],
	)
	if not rows:
		return None

	target_hours = flt(hours)
	best = None
	best_delta = None
	for row in rows:
		row_hours = flt(row.custom_estimated_timehours)
		delta = abs(row_hours - target_hours) if target_hours else 0
		if best is None or delta < best_delta:
			best = row.name
			best_delta = delta
	return best


def _build_package_parts(
	rows: list[tuple], col_idx: int, parts_start: int | None, part_catalog: dict
) -> list[dict]:
	if parts_start is None:
		return []

	parts: list[dict] = []
	seen: set[str] = set()

	# Oil / consumables above the Parts header
	for label in ("oil", "other consumables"):
		row_idx = next((i for i, row in enumerate(rows) if _norm_label(row[0]) == label), None)
		if row_idx is None:
			continue
		amount = _cell_number(rows[row_idx][col_idx])
		if amount <= 0:
			continue
		part_name = _ensure_package_part(label.title(), amount=amount, catalog=part_catalog)
		if part_name and part_name not in seen:
			seen.add(part_name)
			parts.append({"part_item": part_name, "quantity": 1, "unit_price": amount})

	for row_idx in range(parts_start + 1, len(rows)):
		label = _text(rows[row_idx][0])
		norm = _norm_label(label)
		if not label or norm in PART_SECTION_STOP:
			break
		if norm in LABOUR_META_LABELS:
			break

		amount = _cell_number(rows[row_idx][col_idx])
		if amount <= 0:
			continue

		part_name = _ensure_package_part(label, amount=amount, catalog=part_catalog)
		if part_name and part_name not in seen:
			seen.add(part_name)
			parts.append({"part_item": part_name, "quantity": 1, "unit_price": amount})

	return parts


def _ensure_package_part(label: str, amount: float = 0, catalog: dict | None = None) -> str | None:
	label = (label or "").strip()
	if not label:
		return None

	catalog = catalog or {}
	catalog_hit = catalog.get(_norm_label(label))
	oem = (catalog_hit or {}).get("part_no") or _slug_part_code(label)
	description = (catalog_hit or {}).get("description") or label
	price = flt((catalog_hit or {}).get("price")) or flt(amount)

	if frappe.db.exists("Spare Part", oem):
		if price:
			frappe.db.set_value("Spare Part", oem, "selling_price", price, update_modified=False)
		return oem

	existing = frappe.db.get_value("Spare Part", {"item_name": ["like", f"%{description[:40]}%"]}, "name")
	if existing:
		return existing

	item_group = (
		frappe.db.get_single_value("DMS Settings", "default_item_group") or "All Item Groups"
	)
	if not frappe.db.exists("Item", oem):
		frappe.get_doc(
			{
				"doctype": "Item",
				"item_code": oem,
				"item_name": description[:140],
				"item_group": item_group,
				"stock_uom": "Nos",
				"is_stock_item": 1,
				"is_sales_item": 1,
				"is_purchase_item": 1,
				"standard_rate": price,
			}
		).insert(ignore_permissions=True)

	if not frappe.db.exists("Spare Part", oem):
		frappe.get_doc(
			{
				"doctype": "Spare Part",
				"spare_part_item": oem,
				"oem_part_number": oem,
				"item_name": description[:140],
				"selling_price": price or None,
				"part_category": "Genuine Part",
			}
		).insert(ignore_permissions=True)

	return oem


def _build_part_catalog(rows: list[tuple]) -> dict:
	catalog: dict[str, dict] = {}
	for row in rows:
		part_no = _text(row[0] if len(row) > 0 else "")
		description = _text(row[1] if len(row) > 1 else "")
		price = _cell_number(row[4] if len(row) > 4 else None)
		if not description and not part_no:
			continue
		if _norm_label(part_no) == "part no":
			continue

		entry = {
			"part_no": part_no or _slug_part_code(description),
			"description": description or part_no,
			"price": price,
		}
		for key in filter(None, [_norm_label(description), _norm_label(part_no)]):
			catalog.setdefault(key, entry)

		# Loose aliases for grouped labels like "Oil Filter + Washer"
		if description:
			for token in re.split(r"[+/]", description):
				token = _norm_label(token)
				if token:
					catalog.setdefault(token, entry)
	return catalog


def _is_package_column(package_id: str) -> bool:
	text = (package_id or "").strip()
	if not text or text.lower() in {"disct", "discount"}:
		return False
	return bool(re.match(r"^\d", text))


def _parse_package_columns(header_row: tuple) -> list[tuple[int, str, int | None]]:
	packages: list[tuple[int, str, int | None]] = []
	seen_ids: set[str] = set()
	for col_idx, raw in enumerate(header_row):
		if col_idx < 2:
			continue
		package_id = _text(raw)
		if not _is_package_column(package_id):
			continue
		pkg_key = re.sub(r"[^\w]+", "", package_id.upper())
		if pkg_key in seen_ids:
			continue
		seen_ids.add(pkg_key)
		packages.append((col_idx, package_id, _package_id_to_km(package_id)))
	return packages


def _parse_summary_intervals(wb) -> dict[int, dict]:
	if "SUMMARY" not in wb.sheetnames:
		return {}

	ws = wb["SUMMARY"]
	out: dict[int, dict] = {}
	for row in ws.iter_rows(values_only=True):
		label = _text(row[0] if row else "")
		if not label or "km" not in label.lower():
			continue
		km, months = _parse_interval_label(label)
		if km:
			out[km] = {"months": months, "label": label.replace("\n", " ").strip()}
	return out


def _parse_interval_label(label: str) -> tuple[int | None, int | None]:
	km_match = re.search(r"([\d,]+)\s*KM", label, flags=re.I)
	month_match = re.search(r"([\d,]+)\s*Months?", label, flags=re.I)
	km = cint(km_match.group(1).replace(",", "")) if km_match else None
	months = cint(month_match.group(1).replace(",", "")) if month_match else None
	return km, months


def _package_id_to_km(package_id: str) -> int | None:
	text = (package_id or "").strip().lower()
	match = re.match(r"(\d+)\s*k", text)
	if match:
		return cint(match.group(1)) * 1000
	return None


def _km_to_pm_sub_code(interval_km: int | None, hours: float) -> int | None:
	if not interval_km:
		return None
	km = cint(interval_km)
	if km % 40000 == 0:
		return 40
	if km % 20000 == 0:
		return 20
	if km % 10000 == 0:
		return 10
	return 5


def _package_name(model_code: str, package_id: str) -> str:
	model_code = (model_code or "MODEL").strip().upper()
	pkg = re.sub(r"[^\w]+", "", (package_id or "").upper()) or "PKG"
	return f"{model_code}-{pkg}"


def _package_description(
	model_name: str, package_id: str, interval_km: int | None, interval_months: int | None
) -> str:
	parts = [model_name, package_id]
	if interval_km:
		parts.append(f"{interval_km:,} KM")
	if interval_months:
		parts.append(f"{interval_months} Months")
	return " · ".join(p for p in parts if p)


def _known_sheet_model(sheet_key: str) -> tuple[str, str] | None:
	return _KNOWN_SHEET_MODELS.get(sheet_key)


def _model_lookup_keys(model_name: str, model_code: str) -> frozenset[str]:
	return frozenset({_norm_key(model_name), _norm_key(model_code)}) - {""}


def _row_value(row, field: str):
	if isinstance(row, dict):
		return row.get(field)
	return getattr(row, field, None)


def _find_vehicle_model_by_code(
	models: list[dict], model_code: str
) -> tuple[str, str, str] | None:
	code_key = _norm_key(model_code)
	for row in models:
		if _norm_key(_row_value(row, "model_code")) == code_key:
			return _row_value(row, "model_name"), _row_value(row, "model_code"), _row_value(row, "name")
	return None


def _find_vehicle_model_row(
	models: list[dict], lookup_keys: frozenset[str]
) -> tuple[str, str, str] | None:
	for row in models:
		row_keys = _model_lookup_keys(_row_value(row, "model_name"), _row_value(row, "model_code"))
		if row_keys & lookup_keys:
			return _row_value(row, "model_name"), _row_value(row, "model_code"), _row_value(row, "name")
	return None


def _resolve_model_for_key(
	sheet_key: str, models: list[dict]
) -> tuple[str, str, str, int] | None:
	if not sheet_key:
		return None

	known = _known_sheet_model(sheet_key)
	if known:
		model_name, model_code = known
		matched = _find_vehicle_model_by_code(models, model_code)
		if matched:
			return *matched, 0

		brand = ensure_brand(DEFAULT_BRAND)
		existed = frappe.db.exists("Vehicle Model", {"model_code": model_code})
		vehicle_model = ensure_vehicle_model(
			model_code=model_code, model_name=model_name, brand=brand
		)
		return model_name, model_code, vehicle_model, 0 if existed else 1

	matched = _find_vehicle_model_by_code(models, sheet_key)
	if matched:
		return *matched, 0

	for row in models:
		row_keys = _model_lookup_keys(_row_value(row, "model_name"), _row_value(row, "model_code"))
		if sheet_key in row_keys:
			return _row_value(row, "model_name"), _row_value(row, "model_code"), _row_value(row, "name"), 0

	return None


def _resolve_vehicle_model(
	sheet_name: str, sheet_label: str
) -> tuple[str, str, str, int] | None:
	models = frappe.get_all(
		"Vehicle Model", fields=["name", "model_name", "model_code"], filters={"is_active": 1}
	)

	sheet_key = _norm_key(sheet_name)
	resolved = _resolve_model_for_key(sheet_key, models)
	if resolved:
		return resolved

	label_key = _norm_key(sheet_label)
	if label_key and label_key != sheet_key:
		resolved = _resolve_model_for_key(label_key, models)
		if resolved:
			return resolved

	search_keys = {sheet_key, label_key} - {""}
	if any("DASHING" in key for key in search_keys):
		for row in models:
			model_name = _row_value(row, "model_name")
			model_code = _row_value(row, "model_code")
			if "DASHING" in _norm_key(model_name) or _norm_key(model_code) == "JD15":
				return model_name, model_code, _row_value(row, "name"), 0

	return None


def _first_row_index(rows: list[tuple], label: str, before_label: str | None = None) -> int | None:
	target = _norm_label(label)
	stop = _norm_label(before_label) if before_label else None
	for idx, row in enumerate(rows):
		norm = _norm_label(row[0] if row else "")
		if stop and norm == stop:
			break
		if norm == target:
			return idx
	return None


def _norm_label(value) -> str:
	return re.sub(r"\s+", " ", _text(value)).strip().lower()


def _text(value) -> str:
	if value is None:
		return ""
	if isinstance(value, float) and value == int(value):
		return str(int(value))
	return str(value).strip()


def _cell_number(value) -> float:
	if value in (None, "", "#N/A"):
		return 0.0
	try:
		return flt(value)
	except (TypeError, ValueError):
		return 0.0


def _slug_part_code(text: str) -> str:
	base = re.sub(r"[^\w]+", "-", (text or "").upper()).strip("-")
	return (base[:140] if base else "PART")
