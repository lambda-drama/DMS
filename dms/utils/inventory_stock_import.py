"""Import Jetour inventory workbook into Item, Spare Part, Item Price, and Stock Reconciliation."""

from __future__ import annotations

import os
import re
from dataclasses import dataclass

import frappe
from frappe import _
from frappe.utils import cint, flt, get_files_path

from dms.dealer_management_system.utils.stock_operations import create_dms_stock_reconciliation

INVENTORY_WAREHOUSE = "Service Center Addis Ababa - SM"
DEFAULT_STOCK_UOM = "Nos"
INVENTORY_PRICE_CURRENCY = "ETB"
INVENTORY_SELLING_PRICE_LIST = "Ethiopia Local Sales"

_CATEGORY_TO_PART_CATEGORY = {
	"Engine & Cooling": "Engine Parts",
	"Electrical & Electronics": "Electrical & Electronics",
	"Body & Trim": "Body & Interior",
	"General / Trim": "Body & Interior",
	"Suspension & Steering": "Suspension & Steering",
	"Hardware & Fasteners": "Fasteners & Hardware",
	"Drivetrain & Transmission": "Transmission Parts",
	"Braking System": "Brake System",
	"Filter & Maintenance": "Filters",
	"Fluids & Lubricants": "Fluids & Lubricants",
	"Ignition System": "Electrical & Electronics",
	"Fuel System": "Fuel System",
}

_HEADER_ALIASES = {
	"part_no": {"partno", "partnumber", "partno."},
	"part_name": {"partname"},
	"qty": {"qty", "qtypcs", "qty(pcs)", "qtypcs)"},
	"unit_price_exw_etb": {"unitpriceexwetb"},
	"total_unit_price_with_tt": {"totalunitpricewithtt"},
	"retail_price": {"ratailprice", "retailprice"},
	"location": {"location"},
	"category": {"category"},
}


@dataclass(slots=True)
class InventoryRow:
	part_no: str
	part_name: str
	qty: float
	unit_price_exw_etb: float
	landed_cost: float
	retail_price: float
	location: str
	category: str


def import_inventory_stock_workbook(file_path: str) -> dict:
	rows = _read_inventory_rows(file_path)
	summary = {
		"rows_processed": len(rows),
		"rows_skipped": 0,
		"item_groups_created": 0,
		"items_created": 0,
		"items_reused": 0,
		"spare_parts_created": 0,
		"spare_parts_updated": 0,
	}

	created_groups: set[str] = set()
	for row in rows:
		group_name, group_created = _ensure_item_group(row.category)
		if group_created and group_name not in created_groups:
			created_groups.add(group_name)
			summary["item_groups_created"] += 1

		item_created = _upsert_item(row, group_name)
		spare_created = _upsert_spare_part(row)

		summary["items_created" if item_created else "items_reused"] += 1
		summary["spare_parts_created" if spare_created else "spare_parts_updated"] += 1

		if item_created:
			_sync_import_item_price(row)

	frappe.db.commit()
	return summary


def create_inventory_stock_reconciliation_from_workbook(
	file_path: str,
	posting_date: str | None = None,
	submit: int | bool = 1,
	warehouse: str = INVENTORY_WAREHOUSE,
) -> dict:
	rows = _read_inventory_rows(file_path)
	warehouse = (warehouse or INVENTORY_WAREHOUSE).strip()
	company = _warehouse_company(warehouse)

	created_groups: set[str] = set()
	for row in rows:
		group_name, group_created = _ensure_item_group(row.category)
		if group_created:
			created_groups.add(group_name)
		item_created = _upsert_item(row, group_name)
		_upsert_spare_part(row)
		if item_created:
			_sync_import_item_price(row)

	items, duplicate_rows_merged = _build_stock_reconciliation_items(rows)

	result = create_dms_stock_reconciliation(
		{
			"company": company,
			"warehouse": warehouse,
			"posting_date": posting_date or frappe.utils.today(),
			"submit": cint(submit),
			"remarks": _("Inventory stock import from workbook"),
			"items": items,
		}
	)
	frappe.db.commit()
	return {
		**result,
		"warehouse": warehouse,
		"company": company,
		"rows_processed": len(rows),
		"unique_items": len(items),
		"duplicate_rows_merged": duplicate_rows_merged,
		"item_groups_created": len(created_groups),
	}


def _find_etb_selling_price_list() -> str | None:
	return frappe.db.get_value(
		"Price List",
		{"currency": INVENTORY_PRICE_CURRENCY, "enabled": 1, "selling": 1},
		"name",
		order_by="creation asc",
	)


def _get_inventory_selling_price_list() -> str | None:
	from dms.dealer_management_system.utils.stock_operations import get_dms_default_selling_price_list

	return get_dms_default_selling_price_list()


def _default_item_company() -> str:
	try:
		return _warehouse_company(INVENTORY_WAREHOUSE)
	except frappe.ValidationError:
		company = (frappe.defaults.get_global_default("company") or "").strip()
		if company:
			return company
		frappe.throw(_("Set a default Company before importing inventory."))


def _resolve_inventory_price_list(price_list: str | None = None) -> tuple[str, str]:
	price_list = (price_list or "").strip()
	if price_list and not frappe.db.exists("Price List", price_list):
		frappe.throw(_("Price List {0} was not found.").format(frappe.bold(price_list)))

	if price_list:
		currency = frappe.db.get_value("Price List", price_list, "currency")
		if currency == INVENTORY_PRICE_CURRENCY:
			return price_list, INVENTORY_PRICE_CURRENCY
		frappe.throw(
			_(
				"Price List {0} uses {1}. Inventory retail prices are in {2} — select an ETB selling price list."
			).format(
				frappe.bold(price_list),
				frappe.bold(currency or "?"),
				frappe.bold(INVENTORY_PRICE_CURRENCY),
			)
		)

	price_list = _get_inventory_selling_price_list()
	if not price_list:
		frappe.throw(
			_(
				'No enabled ETB selling Price List found. Create "{0}" (currency ETB) or another ETB selling price list.'
			).format(INVENTORY_SELLING_PRICE_LIST)
		)
	return price_list, INVENTORY_PRICE_CURRENCY


def _sync_import_item_price(row: InventoryRow) -> None:
	if flt(row.retail_price) <= 0:
		return
	price_list, currency = _resolve_inventory_price_list(INVENTORY_SELLING_PRICE_LIST)
	_upsert_item_price(row.part_no, price_list, currency, row.retail_price)


def create_inventory_item_prices_from_workbook(file_path: str, price_list: str) -> dict:
	rows = _read_inventory_rows(file_path)
	price_list, currency = _resolve_inventory_price_list(price_list)

	summary = {
		"rows_processed": len(rows),
		"item_groups_created": 0,
		"items_created": 0,
		"items_reused": 0,
		"spare_parts_created": 0,
		"spare_parts_updated": 0,
		"prices_created": 0,
		"prices_updated": 0,
		"prices_skipped": 0,
		"price_list": price_list,
		"currency": currency,
	}

	created_groups: set[str] = set()
	for row in rows:
		group_name, group_created = _ensure_item_group(row.category)
		if group_created and group_name not in created_groups:
			created_groups.add(group_name)
			summary["item_groups_created"] += 1

		item_created = _upsert_item(row, group_name)
		spare_created = _upsert_spare_part(row)
		summary["items_created" if item_created else "items_reused"] += 1
		summary["spare_parts_created" if spare_created else "spare_parts_updated"] += 1

		if flt(row.retail_price) <= 0:
			summary["prices_skipped"] += 1
			continue

		price_created = _upsert_item_price(row.part_no, price_list, currency, row.retail_price)
		summary["prices_created" if price_created else "prices_updated"] += 1

	frappe.db.commit()
	return summary


def _resolve_file_path(file_url: str) -> str:
	file_url = (file_url or "").strip()
	if not file_url:
		frappe.throw(_("File URL is required."))

	if file_url.startswith("/private/files/"):
		return get_files_path(*file_url.replace("/private/files/", "").split("/"), is_private=1)

	if file_url.startswith("/files/"):
		return get_files_path(*file_url.replace("/files/", "").split("/"), is_private=0)

	if os.path.isabs(file_url) and os.path.isfile(file_url):
		return file_url

	frappe.throw(_("Could not resolve uploaded file path."))


def import_inventory_stock_file_url(file_url: str) -> dict:
	return import_inventory_stock_workbook(_resolve_file_path(file_url))


def create_inventory_stock_reconciliation_file_url(
	file_url: str,
	posting_date: str | None = None,
	submit: int | bool = 1,
	warehouse: str = INVENTORY_WAREHOUSE,
) -> dict:
	return create_inventory_stock_reconciliation_from_workbook(
		_resolve_file_path(file_url), posting_date=posting_date, submit=submit, warehouse=warehouse
	)


def create_inventory_item_prices_file_url(file_url: str, price_list: str) -> dict:
	return create_inventory_item_prices_from_workbook(_resolve_file_path(file_url), price_list=price_list)


def _read_inventory_rows(file_path: str) -> list[InventoryRow]:
	try:
		import openpyxl
	except ImportError as exc:
		raise ImportError(_("Excel import requires openpyxl in the bench environment.")) from exc

	if not os.path.isfile(file_path):
		frappe.throw(_("File not found: {0}").format(file_path))

	wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
	try:
		ws = wb[wb.sheetnames[0]]
		raw_rows = list(ws.iter_rows(values_only=True))
	finally:
		wb.close()

	if not raw_rows:
		return []

	header_map = _header_map(raw_rows[0])
	required = ("part_no", "part_name", "qty", "retail_price", "location", "category")
	missing = [field for field in required if field not in header_map]
	if missing:
		frappe.throw(_("Inventory workbook is missing columns: {0}").format(", ".join(missing)))

	rows: list[InventoryRow] = []
	for raw in raw_rows[1:]:
		part_no = _text(raw[header_map["part_no"]] if len(raw) > header_map["part_no"] else None)
		part_name = _text(raw[header_map["part_name"]] if len(raw) > header_map["part_name"] else None)
		if not part_no or not part_name:
			continue

		qty = _number(raw[header_map["qty"]] if len(raw) > header_map["qty"] else None)
		location = _text(raw[header_map["location"]] if len(raw) > header_map["location"] else None)
		category = _text(raw[header_map["category"]] if len(raw) > header_map["category"] else None) or "Other"
		unit_price_exw_etb = _number(
			raw[header_map["unit_price_exw_etb"]] if "unit_price_exw_etb" in header_map and len(raw) > header_map["unit_price_exw_etb"] else None
		)
		landed_cost = _number(
			raw[header_map["total_unit_price_with_tt"]] if "total_unit_price_with_tt" in header_map and len(raw) > header_map["total_unit_price_with_tt"] else None
		)
		retail_price = _number(
			raw[header_map["retail_price"]] if len(raw) > header_map["retail_price"] else None
		)
		rows.append(
			InventoryRow(
				part_no=part_no,
				part_name=part_name[:140],
				qty=qty,
				unit_price_exw_etb=unit_price_exw_etb,
				landed_cost=landed_cost or unit_price_exw_etb,
				retail_price=retail_price,
				location=location,
				category=category[:140],
			)
		)

	return rows


def _header_map(header_row: tuple | list) -> dict[str, int]:
	out: dict[str, int] = {}
	for idx, raw in enumerate(header_row or []):
		key = _norm_header(raw)
		if not key:
			continue
		for fieldname, aliases in _HEADER_ALIASES.items():
			if key in aliases:
				out.setdefault(fieldname, idx)
	return out


def _norm_header(value) -> str:
	return re.sub(r"[^a-z0-9]+", "", _text(value).lower())


def _text(value) -> str:
	if value is None:
		return ""
	if isinstance(value, float) and value == int(value):
		return str(int(value))
	return str(value).strip()


def _number(value) -> float:
	return flt(value or 0)


def _ensure_item_group(category: str) -> tuple[str, bool]:
	category = (category or "").strip() or "Other"
	if frappe.db.exists("Item Group", category):
		return category, False

	parent_group = (frappe.db.get_single_value("DMS Settings", "default_item_group") or "All Item Groups").strip()
	if not frappe.db.exists("Item Group", parent_group):
		frappe.throw(_("Parent Item Group {0} was not found.").format(frappe.bold(parent_group)))

	doc = frappe.get_doc(
		{
			"doctype": "Item Group",
			"item_group_name": category,
			"parent_item_group": parent_group,
			"is_group": 0,
		}
	)
	if frappe.get_meta("Item Group").has_field("custom_auto_generate_spare_parts"):
		doc.custom_auto_generate_spare_parts = 1
	doc.insert(ignore_permissions=True)
	return doc.name, True


def _upsert_item(row: InventoryRow, item_group: str) -> bool:
	if frappe.db.exists("Item", row.part_no):
		return False

	item_data = {
		"item_code": row.part_no,
		"item_name": row.part_name,
		"description": row.part_name,
		"item_group": item_group,
		"stock_uom": DEFAULT_STOCK_UOM,
		"is_stock_item": 1,
		"is_sales_item": 1,
		"is_purchase_item": 1,
		"include_item_in_manufacturing": 0,
		"standard_rate": row.landed_cost or row.unit_price_exw_etb or row.retail_price,
		"disabled": 0,
	}

	if _get_inventory_selling_price_list():
		item_data["item_defaults"] = [
			{
				"company": _default_item_company(),
				"default_price_list": INVENTORY_SELLING_PRICE_LIST,
			}
		]

	doc = frappe.get_doc({"doctype": "Item", **item_data})
	doc.insert(ignore_permissions=True)
	return True


def _upsert_spare_part(row: InventoryRow) -> bool:
	part_category = _CATEGORY_TO_PART_CATEGORY.get(row.category, "Other")
	markup_percentage = _derive_markup_percentage(row.landed_cost, row.retail_price)
	spare_name = (
		frappe.db.get_value("Spare Part", {"spare_part_item": row.part_no}, "name")
		or frappe.db.get_value("Spare Part", {"oem_part_number": row.part_no}, "name")
		or (row.part_no if frappe.db.exists("Spare Part", row.part_no) else None)
	)

	spare_data = {
		"spare_part_item": row.part_no,
		"oem_part_number": row.part_no,
		"manufacturer_part_number": row.part_no,
		"part_category": part_category,
		"part_subcategory": row.category,
		"part_type": "Genuine OEM",
		"bin_location": row.location,
		"last_purchase_price": row.landed_cost or None,
		"markup_percentage": markup_percentage if markup_percentage > 0 else None,
		"selling_price": row.retail_price or None,
	}

	if spare_name:
		doc = frappe.get_doc("Spare Part", spare_name)
		for fieldname, value in spare_data.items():
			setattr(doc, fieldname, value)
		doc.save(ignore_permissions=True)
		return False

	doc = frappe.get_doc({"doctype": "Spare Part", **spare_data})
	doc.insert(ignore_permissions=True)
	return True


def _upsert_item_price(item_code: str, price_list: str, currency: str, rate: float) -> bool:
	filters = {"item_code": item_code, "price_list": price_list, "selling": 1}
	existing = frappe.db.get_value("Item Price", filters, "name")
	price_data = {
		"item_code": item_code,
		"price_list": price_list,
		"price_list_rate": flt(rate),
		"currency": currency,
		"uom": DEFAULT_STOCK_UOM,
		"selling": 1,
	}

	if existing:
		doc = frappe.get_doc("Item Price", existing)
		for fieldname, value in price_data.items():
			setattr(doc, fieldname, value)
		doc.save(ignore_permissions=True)
		return False

	doc = frappe.get_doc({"doctype": "Item Price", **price_data})
	doc.insert(ignore_permissions=True)
	return True


def _derive_markup_percentage(cost: float, retail: float) -> float:
	cost = flt(cost)
	retail = flt(retail)
	if cost <= 0 or retail <= cost:
		return 0.0
	return round(((retail - cost) / cost) * 100, 4)


def _build_stock_reconciliation_items(rows: list[InventoryRow]) -> tuple[list[dict], int]:
	aggregated: dict[str, dict] = {}
	duplicate_rows_merged = 0

	for row in rows:
		item_code = (row.part_no or "").strip()
		if not item_code:
			continue

		qty = max(flt(row.qty), 0)
		rate = flt(row.landed_cost or row.unit_price_exw_etb or row.retail_price)
		total_value = qty * rate

		if item_code in aggregated:
			duplicate_rows_merged += 1
			aggregated[item_code]["qty"] += qty
			aggregated[item_code]["total_value"] += total_value
			continue

		aggregated[item_code] = {
			"item_code": item_code,
			"qty": qty,
			"total_value": total_value,
			"fallback_rate": rate,
		}

	items: list[dict] = []
	for item_code, entry in aggregated.items():
		qty = flt(entry["qty"])
		total_value = flt(entry["total_value"])
		fallback_rate = flt(entry["fallback_rate"])
		valuation_rate = (total_value / qty) if qty > 0 else fallback_rate
		items.append(
			{
				"item_code": item_code,
				"qty": qty,
				"valuation_rate": valuation_rate,
			}
		)

	return items, duplicate_rows_merged


def _warehouse_company(warehouse: str) -> str:
	if not frappe.db.exists("Warehouse", warehouse):
		frappe.throw(_("Warehouse {0} was not found.").format(frappe.bold(warehouse)))
	company = (frappe.db.get_value("Warehouse", warehouse, "company") or "").strip()
	if not company:
		frappe.throw(_("Company is missing on Warehouse {0}.").format(frappe.bold(warehouse)))
	return company
